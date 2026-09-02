#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tự động trích xuất ca kiểm thử từ test_cases.md và kết xuất file Excel test_cases.xlsx chuẩn hóa
Tác giả: Lưu Ngô Quốc Bảo (MSSV: 23127327)
"""

import re
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def parse_markdown_tables(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Tách theo từng API
    sections = re.split(r'# \d+\. API ', content)
    apis_data = {}

    for sec in sections[1:]:
        lines = sec.strip().split('\n')
        api_title = lines[0].strip()
        
        # Tìm tất cả các dòng bảng | ... |
        table_rows = []
        for line in lines:
            line_str = line.strip()
            if line_str.startswith('|') and line_str.endswith('|'):
                # Bỏ qua dòng header separator | :--- | :--- |
                if '---' in line_str:
                    continue
                cells = [c.strip().replace('**', '') for c in line_str.split('|')[1:-1]]
                if cells:
                    table_rows.append(cells)
        
        apis_data[api_title] = table_rows

    return apis_data

def create_excel(apis_data, output_path):
    wb = openpyxl.Workbook()
    
    # Header styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # ----------------------------------------------------
    # Sheet 1: Summary Dashboard
    # ----------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Test Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.cell(row=1, column=1, value="BÁO CÁO TỔNG HỢP KIỂM THỬ API - ESHOP SUT").font = title_font
    ws_sum.cell(row=2, column=1, value="Sinh viên: Lưu Ngô Quốc Bảo | MSSV: 23127327 | Môn: Kiểm thử phần mềm").font = bold_font
    
    summary_headers = ["Phân Hệ / API", "Endpoint", "AI Tests", "Human Audit", "Human Ext", "Tổng Tests", "Requests", "Assertions", "Passed", "Failed", "Số Bugs"]
    for col_idx, h in enumerate(summary_headers, 1):
        c = ws_sum.cell(row=4, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    
    summary_rows = [
        ["Health Check", "GET /api/products", 0, 0, 1, 1, 1, 3, 3, 0, 0],
        ["Pool A: FR-02", "POST /api/login", 38, 38, 6, 44, 46, 64, 60, 4, 4],
        ["Pool B: FR-08", "POST /api/checkout", 36, 36, 7, 43, 47, 58, 41, 17, 5],
        ["Pool C: FR-14", "CRUD /api/categories", 38, 38, 6, 44, 48, 54, 35, 19, 4],
        ["TỔNG CỘNG", "Toàn Bộ Hệ Thống", 112, 112, 20, 132, 142, 179, 139, 40, 13]
    ]
    
    for r_idx, row in enumerate(summary_rows, 5):
        for c_idx, val in enumerate(row, 1):
            cell = ws_sum.cell(row=r_idx, column=c_idx, value=val)
            cell.font = bold_font if r_idx == 9 else regular_font
            cell.border = thin_border
            if c_idx >= 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_idx == 9:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ----------------------------------------------------
    # Sheets for each API
    # ----------------------------------------------------
    sheet_names = ["FR-02 Login", "FR-08 Checkout", "FR-14 Categories"]
    api_keys = list(apis_data.keys())

    for idx, name in enumerate(sheet_names):
        ws = wb.create_sheet(title=name)
        ws.views.sheetView[0].showGridLines = True
        
        if idx < len(api_keys):
            data = apis_data[api_keys[idx]]
            headers = ["Test ID", "Phân loại", "Tên ca kiểm thử", "Dữ liệu đầu vào (Payload / URL)", "Kỳ vọng theo Đặc tả (Expected Result)", "Thẩm định", "Lý do & Hiệu chỉnh"]
            
            # Ghi header
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col_idx, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Ghi dữ liệu
            row_cursor = 2
            for row in data:
                # Bỏ qua dòng header nếu trùng
                if len(row) >= 5 and ('Test ID' in row[0] or 'ID' in row[0]):
                    continue
                if len(row) < 5:
                    continue
                
                for col_idx, val in enumerate(row[:7], 1):
                    cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                    cell.font = regular_font
                    cell.border = thin_border
                    if col_idx in [1, 2, 6]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                row_cursor += 1

    # ----------------------------------------------------
    # Sheet 5: 13 Verified Bugs
    # ----------------------------------------------------
    ws_bugs = wb.create_sheet(title="Bug Inventory")
    ws_bugs.views.sheetView[0].showGridLines = True
    
    bug_headers = ["Bug ID", "API Bị Ảnh Hưởng", "Mức Độ (Severity)", "Tên Lỗi Kỹ Thuật", "Vị Trí Mã Nguồn (Root Cause)", "Hành Vi Lỗi (Actual vs Expected)"]
    for col_idx, h in enumerate(bug_headers, 1):
        c = ws_bugs.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    
    bugs = [
        ["BUG-01", "POST /api/login", "Major", "Bộ đếm đăng nhập sai tăng sai bước nhảy (+2 thay vì +1)", "server.js:54", "Khóa tài khoản sau 2 lần sai thay vì 3 lần"],
        ["BUG-02", "POST /api/login", "Medium", "Thời gian khóa tài khoản sai (180s thay vì 30s)", "server.js:57", "Khóa 180,000ms thay vì 30,000ms theo đặc tả demo"],
        ["BUG-03", "POST /api/login", "Critical", "Để lộ Plaintext Password trong response (SEC-01)", "server.js:35", "Trả về nguyên vẹn user.password trong body JSON"],
        ["BUG-04", "POST /api/login", "Medium", "Email phân biệt hoa thường sai chuẩn RFC 5321", "server.js:35", "Không dùng LOWER(email), từ chối TEST@ESHOP.COM"],
        ["BUG-05", "POST /api/checkout", "Critical", "Lỗ hổng gian lận giá Price Tampering nghiêm trọng", "server.js:297", "Tin tưởng total_amount từ client, tạo đơn hàng 0 VND"],
        ["BUG-06", "POST /api/checkout", "Major", "Giỏ hàng không được làm rỗng sau khi checkout", "server.js:305", "Thiếu lệnh xóa userCarts[userId] = []"],
        ["BUG-07", "POST /api/checkout", "Major", "Cho phép tạo đơn hàng khi giỏ hàng rỗng", "server.js:297", "Không kiểm tra độ dài giỏ hàng trước khi insert order"],
        ["BUG-08", "POST /api/checkout", "Major", "Thiếu hoàn toàn Validation trên Checkout", "server.js:297", "Chấp nhận tiền âm, bằng 0, địa chỉ rỗng, null"],
        ["BUG-09", "POST /api/checkout", "Major", "Lỗ hổng Overselling & Tồn kho âm khi tương tranh", "database.js / server.js:297", "Không có khóa giao dịch locking, 2 người mua món hàng cuối cùng"],
        ["BUG-10", "POST/PUT/DELETE /api/categories", "Critical", "Lỗ hổng BFLA trên các Endpoint Danh mục (SEC-03)", "server.js:249, 257, 269", "Thiếu kiểm tra req.user.role === 'admin', user thường xóa được"],
        ["BUG-11", "POST/PUT /api/categories", "Major", "Thiếu Validation tên danh mục", "server.js:251", "Chấp nhận tên rỗng, null, khoảng trắng"],
        ["BUG-12", "PUT/DELETE /api/categories/:id", "Medium", "Vi phạm chuẩn RESTful 404 khi ID không tồn tại", "server.js:263, 274", "Trả về 200 OK khi ID = 999999 do không check this.changes"],
        ["BUG-13", "DELETE /api/categories/:id", "Major", "Vi phạm toàn vẹn quan hệ khi xóa danh mục có sản phẩm", "server.js:271", "Cho phép xóa danh mục đang chứa sản phẩm, gây mồ côi dữ liệu"]
    ]
    
    for r_idx, b in enumerate(bugs, 2):
        for c_idx, val in enumerate(b, 1):
            cell = ws_bugs.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if c_idx in [1, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Tự căn chỉnh độ rộng cột cho tất cả các sheet
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len and len(val_str) < 60:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_path)
    print("[Success] Generated test_cases.xlsx successfully!")


if __name__ == '__main__':
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    test_cases_md = os.path.join(project_root, 'test_cases.md')
    excel_out = os.path.join(project_root, 'test_cases.xlsx')
    
    data = parse_markdown_tables(test_cases_md)
    create_excel(data, excel_out)
